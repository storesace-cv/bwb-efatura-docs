#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_supplier_invoices.py
---------------------------
Exportação de documentos de compras (DFE) do portal eFatura (Cabo Verde) para Excel, com 1 linha por item.

Objetivo
- Listar DFEs (documentos fiscais eletrónicos) num intervalo de datas e descarregar o XML de cada documento.
- Extrair cabeçalhos do documento (fornecedor, datas, nº documento, tipo) e todas as linhas de itens.
- Escrever para Excel com repetição do UID por item (1 item = 1 linha no Excel).

Princípios/invariantes (importante para manutenção)
1) UID é a chave do documento. Um UID pode gerar N linhas no Excel (N = nº de items).
2) O Excel é um "data lake" de linhas: o UID repete-se por item; nunca se deduplica ao nível de linhas.
3) Retoma segura: se houver crash "a meio de um documento", na próxima execução apaga-se tudo desse UID e reescreve-se
   o documento inteiro (opção 2 aprovada).
4) Robustez face ao portal:
   - Paginação pode entrar em loop (ex.: páginas repetidas). Detectamos assinatura repetida e paramos.
   - O XML devolvido pode vir tecnicamente inválido (ex.: '&' não escapado). Há sanitização "best effort".

Comportamento por defeito
- Se o UID já existir no Excel, o script não o volta a descarregar (para ser rápido).
- Exceções:
  - Se existir um UID "in-progress" no ficheiro resume (crash anterior), esse UID é reescrito (apagado+regerado).
  - Se for passado --rewrite-existing, qualquer UID existente será reescrito (atenção: pode ser pesado).

Requisitos
  pip install requests openpyxl

Ficheiros gerados
- Excel: supplier_invoices.xlsx (configurável no INI)
- Estado de retoma: supplier_invoices.xlsx.resume.json (ao lado do Excel)
- Logs: logs/update_supplier_invoices.log (configurável)
- Dumps de respostas problemáticas: logs/bad_responses/ e logs/no_lines/ (quando aplicável)

"""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import html
import json
import logging
import os
import re
import socket
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple, Sequence

import requests
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.efatura_auth import EfaturaAuthManager
from core.exceptions import EfaturaAuthNeedsReauth

SERVICES_BASE = "https://services.efatura.cv"
IAM_BASE = "https://iam.efatura.cv"
DFE_LIST_ENDPOINT = f"{SERVICES_BASE}/v1/dfe"
DFE_XML_ENDPOINT_TMPL = f"{SERVICES_BASE}/v1/dfe/xml/{{uid}}"
USERINFO_ENDPOINT = f"{IAM_BASE}/auth/realms/taxpayers/protocol/openid-connect/userinfo"

BAD_RESPONSE_DIR: Optional[Path] = None

NS_DFE = {"d": "urn:cv:efatura:xsd:v1.0"}


# Excel schema
# v2: inserted "Tipo de Documento" between "Data Documento" and "Numero Documento"
COLUMNS_V1 = [
    "UID",
    "Erro",
    "Nome Fornecedor",
    "NIF Fornecedor",
    "Morada Fornecedor",
    "Data eFatura",
    "Data Documento",
    "Numero Documento",
    "Código Artigo",
    "Nome Artigo",
    "Quantidade",
    "Unidade Medida",
    "Preço Unitário",
    "Desconto",
    "Preço Total (linha)",
    "last_updated",
    "Exported",
]


# =============================================================================
# Excel schema (column order is part of the contract)
# =============================================================================

COLUMNS = [
    "UID",
    "Erro",
    "Nome Fornecedor",
    "NIF Fornecedor",
    "Morada Fornecedor",
    "Data eFatura",
    "Data Documento",
    "Tipo de Documento",
    "Numero Documento",
    "Código Artigo",
    "Nome Artigo",
    "Quantidade",
    "Unidade Medida",
    "Preço Unitário",
    "Desconto",
    "Preço Total (linha)",
    "last_updated",
    "Exported",
]


UID_RE = re.compile(r"^[A-Z]{2}\d{10,}$")



# =============================================================================
# Logging utilities
# =============================================================================

def now_local_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")


LOGGER: Optional[logging.Logger] = None

def setup_logging(log_file: Path) -> None:
    """Configure console + file logging."""
    global LOGGER
    log_file.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("update_supplier_invoices")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    fmt = logging.Formatter("[%(asctime)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    # Reset handlers (re-runs / tests)
    logger.handlers = []

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(str(log_file), encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    LOGGER = logger


def log(msg: str) -> None:
    if LOGGER is None:
        ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{ts}] {msg}", flush=True)
        return
    LOGGER.info(msg)


def log_exception(msg: str) -> None:
    if LOGGER is None:
        log(msg)
        return
    LOGGER.exception(msg)


# =============================================================================
# XML helpers (namespace-agnostic)
# =============================================================================

def get_text(el: Optional[ET.Element]) -> str:
    """Safe text extraction from an ElementTree node."""
    if el is None:
        return ""
    try:
        txt = "".join(el.itertext())
    except Exception:
        txt = el.text or ""
    return (txt or "").strip()


def _localname(tag: str) -> str:
    """Return localname of an XML tag, ignoring namespace."""
    if not tag:
        return ""
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _find_first_by_localnames(root: ET.Element, localnames: Sequence[str]) -> Optional[ET.Element]:
    want = {n.lower() for n in localnames}
    for el in root.iter():
        if _localname(el.tag).lower() in want:
            return el
    return None


def _find_all_by_localnames(root: ET.Element, localnames: Sequence[str]) -> List[ET.Element]:
    want = {n.lower() for n in localnames}
    out: List[ET.Element] = []
    for el in root.iter():
        if _localname(el.tag).lower() in want:
            out.append(el)
    return out


def _find_reference_uids(root: ET.Element) -> List[str]:
    """Heuristic: collect UID-like strings from common reference nodes."""
    refs: List[str] = []
    for el in root.iter():
        ln = _localname(el.tag).lower()
        if "fiscaldocument" in ln or "reference" in ln or "documentreference" in ln or ln.endswith("document"):
            txt = (el.text or "").strip()
            if txt and UID_RE.match(txt):
                refs.append(txt)
    # de-dup while preserving order
    seen = set()
    out: List[str] = []
    for r in refs:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def _coalesce(*vals: str) -> str:
    for v in vals:
        if v:
            return v
    return ""


def _text_anywhere(root: ET.Element, localnames: Sequence[str]) -> str:
    el = _find_first_by_localnames(root, localnames)
    return get_text(el)


_INVALID_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


# =============================================================================
# XML sanitization (defensive; eFatura may emit invalid XML)
# =============================================================================

def sanitize_xml_text(s: str) -> str:
    """Best-effort cleanup for malformed XML returned by eFatura.

    Observed issues in the field:
    - illegal control chars (XML 1.0)
    - bare ampersands in text nodes (e.g. supplier names like "FOOD & EVENTS")

    This function:
    - trims leading garbage before the first '<'
    - removes illegal control chars (except \t \n \r)
    - escapes bare ampersands that are not already part of an entity
    """
    if not s:
        return s

    # Trim leading garbage before first tag
    lt = s.find("<")
    if lt > 0:
        s = s[lt:]

    # Remove illegal chars (except \t \n \r)
    s = _INVALID_XML_CHARS_RE.sub("", s)

    # Escape bare ampersands: & -> &amp; unless already an entity (&amp; &#123; &#x1A; &name;)
    s = re.sub(r"&(?!(?:#\d+|#x[0-9a-fA-F]+|\w+);)", "&amp;", s)

    return s



def dump_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8", errors="replace")



def _atomic_write_json(path: Path, obj: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, path)


def load_resume_state(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        # If corrupted (e.g., crash during write), ignore and start fresh.
        return {}


def save_resume_state(path: Path, state: dict) -> None:
    state = dict(state or {})
    state["ts"] = now_local_iso()
    _atomic_write_json(path, state)


def compute_resume_uid(state: dict) -> Optional[str]:
    """Return a UID that should be rewritten on resume, if we crashed mid-document."""
    started = (state or {}).get("started_uid")
    completed = (state or {}).get("completed_uid")
    if isinstance(started, str) and started and started != completed:
        return started
    return None



def safe_parse_xml(xml_text: str, *, uid: str, stage: str, dump_dir: Path) -> ET.Element:
    """Parse XML with sanitization + dump on failure."""
    try:
        return ET.fromstring(xml_text.encode("utf-8", errors="strict"))
    except Exception as e1:
        cleaned = sanitize_xml_text(xml_text)
        try:
            return ET.fromstring(cleaned.encode("utf-8", errors="strict"))
        except Exception as e2:
            # dump for later forensic analysis
            dump_path = dump_dir / f"{uid}.{stage}.xml"
            dump_text(dump_path, cleaned)
            raise RuntimeError(f"XML_PARSE_ERROR stage={stage} uid={uid} -> dumped to {dump_path}") from e2


def parse_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s.strip())


def safe_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def decode_jwt_exp_unverified(token: str) -> Optional[int]:
    """
    Decode JWT 'exp' claim without verifying signature.
    Returns exp (unix seconds) or None.
    """
    try:
        parts = token.split(".")
        if len(parts) < 2:
            return None
        import base64

        payload_b64 = parts[1]
        padding = "=" * (-len(payload_b64) % 4)
        payload = base64.urlsafe_b64decode(payload_b64 + padding)
        obj = json.loads(payload.decode("utf-8", errors="replace"))
        exp = obj.get("exp")
        return int(exp) if exp is not None else None
    except Exception:
        return None


def resolve_or_fail(hostname: str) -> List[str]:
    try:
        infos = socket.getaddrinfo(hostname, 443, type=socket.SOCK_STREAM)
        ips = sorted({info[4][0] for info in infos})
        return ips
    except Exception as e:
        raise RuntimeError(f"DNS resolution failed for {hostname}: {e}") from e


@dataclass
class Config:
    base_dir: Path
    excel_path: Path
    token_json: Path
    auth_issuer_url: str
    auth_client_id: str
    auth_redirect_uri: str
    auth_scopes: List[str]
    auth_token_store: Path
    auth_client_secret: Optional[str]
    repo_code: str
    date_start: dt.date
    date_end: dt.date
    page_size: int
    timeout_sec: int
    retries: int
    backoff_sec: float
    progress_every: int
    save_every_docs: int
    save_every_seconds: int
    log_file: Path
    verbose: bool


def load_config(path: Path, verbose: bool) -> Config:
    if not path.exists():
        raise FileNotFoundError(f"INI not found: {path}")
    cp = configparser.ConfigParser()
    cp.read(path, encoding="utf-8")

    if "paths" not in cp:
        raise ValueError("INI missing [paths] section")
    if "efatura" not in cp:
        raise ValueError("INI missing [efatura] section")
    if "efatura_auth" not in cp:
        raise ValueError("INI missing [efatura_auth] section")

    base_dir = Path(cp.get("paths", "base_dir", fallback=".")).expanduser().resolve()
    excel_rel = cp.get("paths", "excel_path", fallback="supplier_invoices.xlsx").strip()
    excel_path = Path(excel_rel)
    if not excel_path.is_absolute():
        excel_path = (base_dir / excel_path).resolve()

    token_json = Path(cp.get("efatura", "token_json", fallback="token.json")).expanduser()
    if not token_json.is_absolute():
        token_json = (Path.cwd() / token_json).resolve()

    repo_code = cp.get("efatura", "repo_code", fallback="1").strip()

    issuer_url = cp.get("efatura_auth", "issuer_url", fallback="https://iam.efatura.cv/auth/realms/taxpayers").strip()
    client_id = cp.get("efatura_auth", "client_id", fallback="").strip()
    redirect_uri = cp.get("efatura_auth", "redirect_uri", fallback="").strip()
    scopes_raw = cp.get("efatura_auth", "scopes", fallback="openid profile email offline_access").strip()
    scopes = [s for s in scopes_raw.split() if s]
    token_store = Path(cp.get("efatura_auth", "token_store", fallback="~/.bwb-app/efatura_tokens.json")).expanduser()
    if not token_store.is_absolute():
        token_store = (Path.cwd() / token_store).resolve()
    client_secret = cp.get("efatura_auth", "client_secret", fallback="").strip() or None

    if not client_id:
        raise ValueError("INI missing efatura_auth.client_id")
    if not redirect_uri:
        raise ValueError("INI missing efatura_auth.redirect_uri")

    date_start = parse_date(cp.get("efatura", "date_start"))
    date_end = parse_date(cp.get("efatura", "date_end"))

    page_size = cp.getint("efatura", "page_size", fallback=200)
    timeout_sec = cp.getint("efatura", "timeout_sec", fallback=45)
    retries = cp.getint("efatura", "retries", fallback=3)
    backoff_sec = float(cp.get("efatura", "retry_backoff_sec", fallback="1.5"))

    progress_every = cp.getint("logging", "progress_every_docs", fallback=10) if "logging" in cp else 10
    save_every_docs = cp.getint("logging", "save_every_docs", fallback=progress_every) if "logging" in cp else progress_every
    save_every_seconds = cp.getint("logging", "save_every_seconds", fallback=60) if "logging" in cp else 60

    # Optional file logger path (if relative, resolve under base_dir)
    log_file_cfg = cp.get("logging", "log_file", fallback="").strip() if "logging" in cp else ""
    if log_file_cfg:
        log_file = Path(log_file_cfg).expanduser()
        if not log_file.is_absolute():
            log_file = (base_dir / log_file).resolve()
    else:
        log_file = (base_dir / "logs" / "update_supplier_invoices.log").resolve()


    return Config(
        base_dir=base_dir,
        excel_path=excel_path,
        token_json=token_json,
        auth_issuer_url=issuer_url,
        auth_client_id=client_id,
        auth_redirect_uri=redirect_uri,
        auth_scopes=scopes,
        auth_token_store=token_store,
        auth_client_secret=client_secret,
        repo_code=repo_code,
        date_start=date_start,
        date_end=date_end,
        page_size=page_size,
        timeout_sec=timeout_sec,
        retries=retries,
        backoff_sec=backoff_sec,
        progress_every=progress_every,
        save_every_docs=save_every_docs,
        save_every_seconds=save_every_seconds,
        log_file=log_file,
        verbose=verbose,
    )



# =============================================================================
# eFatura HTTP client
# =============================================================================

class EfaturaClient:
    def __init__(self, access_token_provider: Callable[[], str], repo_code: str, timeout_sec: int, retries: int, backoff_sec: float, verbose: bool):
        self.access_token_provider = access_token_provider
        self.repo_code = repo_code
        self.timeout_sec = timeout_sec
        self.retries = retries
        self.backoff_sec = backoff_sec
        self.verbose = verbose
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "bwb-export/1.0"})

    def _headers(self, accept: str) -> Dict[str, str]:
        return {
            "Authorization": f"Bearer {self.access_token_provider()}",
            "cv-ef-repository-code": str(self.repo_code),
            "Accept": accept,
        }

    def _request(self, method: str, url: str, *, headers: Dict[str, str], params: Optional[Dict[str, Any]] = None) -> requests.Response:
        last_exc: Optional[Exception] = None
        retried_auth = False
        attempts = 0
        while attempts < self.retries:
            attempts += 1
            try:
                if self.verbose:
                    log(f"{method} {url} params={params or {}} (attempt {attempts}/{self.retries})")
                r = self.session.request(method, url, headers=headers, params=params, timeout=self.timeout_sec)
                if r.status_code in (401, 403):
                    # Be explicit: most frequent root cause is expired token
                    body = (r.text or "").lower()
                    if "expired" in body or "invalid_token" in body or "token" in body:
                        raise PermissionError(f"TOKEN_EXPIRED_OR_INVALID (HTTP {r.status_code})")
                return r
            except PermissionError:
                if not retried_auth:
                    retried_auth = True
                    headers = self._headers(headers.get("Accept", "application/json"))
                    continue
                raise
            except requests.exceptions.Timeout as e:
                last_exc = e
                log(f"WARNING: Timeout calling {url} (attempt {attempts}/{self.retries}). Retrying...")
            except requests.exceptions.ConnectionError as e:
                last_exc = e
                msg = str(e)
                if "Could not resolve host" in msg or "Name or service not known" in msg or "nodename nor servname provided" in msg:
                    log(f"WARNING: DNS/Connection error calling {url} (attempt {attempts}/{self.retries}): {e}")
                else:
                    log(f"WARNING: Connection error calling {url} (attempt {attempts}/{self.retries}): {e}")
            except requests.exceptions.SSLError as e:
                last_exc = e
                log(f"WARNING: SSL error calling {url} (attempt {attempts}/{self.retries}): {e}")
            sleep_s = self.backoff_sec * attempts
            time.sleep(sleep_s)
        raise RuntimeError(f"Failed to call {url} after {self.retries} attempts: {last_exc}")

    def userinfo_taxid(self) -> str:
        headers = {"Authorization": f"Bearer {self.access_token_provider()}", "Accept": "application/json"}
        r = self._request("GET", USERINFO_ENDPOINT, headers=headers)
        if r.status_code != 200:
            raise RuntimeError(f"userinfo failed HTTP {r.status_code}: {r.text[:300]}")
        obj = r.json()
        # best-effort: common fields
        for k in ("taxId", "tax_id", "preferred_username", "username", "sub"):
            v = obj.get(k)
            if v:
                return str(v)
        return "OK"

    def list_dfes(self, date_start: dt.date, date_end: dt.date, page_size: int, *, show_fields: bool = False) -> Tuple[List[Dict[str, Any]], List[str]]:
            """
            Returns (records, discovered_date_keys).
            Each record is the raw JSON dict for the DFE listing item.

            Notes:
            - Some eFatura deployments appear to ignore PageSize and/or repeat pages.
              Therefore we stop based on "no new UIDs" and/or repeated page signatures,
              not only on `len(items) < page_size`.
            """
            page = 1
            records: List[Dict[str, Any]] = []
            discovered_date_keys: List[str] = []

            seen_uids: set[str] = set()
            seen_page_sigs: set[tuple] = set()

            def is_last_page(obj: Any) -> Optional[bool]:
                if not isinstance(obj, dict):
                    return None
                for k in ("last", "isLast"):
                    v = obj.get(k)
                    if isinstance(v, bool):
                        return v
                for k in ("hasNext", "hasMore", "has_next", "has_more"):
                    v = obj.get(k)
                    if isinstance(v, bool):
                        return not v
                tp = obj.get("totalPages") or obj.get("total_pages") or obj.get("pages")
                pn = obj.get("page") or obj.get("pageNumber") or obj.get("page_number") or obj.get("number")
                if isinstance(tp, int) and isinstance(pn, int) and tp > 0:
                    if pn >= tp or (pn + 1) >= tp:
                        return True
                return None

            while True:
                params = {
                    "AuthorizedDateStart": date_start.isoformat(),
                    "AuthorizedDateEnd": date_end.isoformat(),
                    "PageSize": int(page_size),
                    "Page": int(page),
                }
                log(f"Fetching listing page {page}...")
                r = self._request("GET", DFE_LIST_ENDPOINT, headers=self._headers("application/json"), params=params)
                if r.status_code != 200:
                    raise RuntimeError(f"List DFEs failed HTTP {r.status_code}: {r.text[:500]}")
                try:
                    obj = r.json()
                except Exception:
                    raise RuntimeError(f"List DFEs did not return JSON. content-type={r.headers.get('content-type')}")

                if show_fields:
                    log("=== LIST RESPONSE TOP-LEVEL KEYS ===")
                    if isinstance(obj, dict):
                        log(", ".join(sorted(obj.keys())))
                    else:
                        log(f"type={type(obj)} (not dict)")
                    first_item = None
                    if isinstance(obj, list) and obj:
                        first_item = obj[0]
                    elif isinstance(obj, dict):
                        for v in obj.values():
                            if isinstance(v, list) and v and isinstance(v[0], dict):
                                first_item = v[0]
                                break
                    if isinstance(first_item, dict):
                        log("=== FIRST ITEM KEYS ===")
                        log(", ".join(sorted(first_item.keys())))
                    else:
                        log("No first item found to inspect.")
                    return [], []

                items = extract_items(obj)
                if not items:
                    log(f"Page {page}: 0 items (stop).")
                    break

                page_uids: List[str] = []
                for it in items:
                    uid = extract_uid_from_item(it)
                    if uid:
                        page_uids.append(uid)
                if page_uids:
                    sig = (page_uids[0], page_uids[-1], len(page_uids))
                else:
                    sig = (str(items[0])[:200], len(items))
                if sig in seen_page_sigs:
                    log(f"Page {page}: repeated page signature (stop; pagination loop suspected). sig={sig}")
                    break
                seen_page_sigs.add(sig)

                for it in items:
                    for k in it.keys():
                        lk = k.lower()
                        if "authorizeddate" in lk or "register" in lk or "created" in lk or "submitted" in lk or "authorization" in lk:
                            if k not in discovered_date_keys:
                                discovered_date_keys.append(k)

                new_items: List[Dict[str, Any]] = []
                for it in items:
                    uid = extract_uid_from_item(it)
                    if uid:
                        if uid in seen_uids:
                            continue
                        seen_uids.add(uid)
                    new_items.append(it)

                if not new_items:
                    log(f"Page {page}: 0 new items after de-dup (stop; pagination loop suspected).")
                    break

                records.extend(new_items)
                log(f"Page {page}: received {len(items)} items, new {len(new_items)} (total unique-ish so far {len(records)}).")

                last_hint = is_last_page(obj)
                if last_hint is True:
                    break
                if len(items) < int(page_size):
                    break

                page += 1
                if page > 10000:
                    raise RuntimeError("Aborting: too many pages (possible pagination loop).")

            return records, discovered_date_keys

    def fetch_dfe_inner_xml(self, uid: str) -> str:
        url = DFE_XML_ENDPOINT_TMPL.format(uid=uid)
        r = self._request("GET", url, headers=self._headers("application/xml"))

        # Basic HTTP validation
        if r.status_code != 200:
            self._dump_http_response(uid, r, stage="http_error", note=f"HTTP {r.status_code}")
            raise RuntimeError(f"Fetch DFE XML failed HTTP {r.status_code}: {r.text[:500]}")

        ct = (r.headers.get("Content-Type") or "").lower()

        # Some failures come back as HTML/JSON while still HTTP 200 (reverse-proxy / WAF / session bounce)
        body_bytes = r.content or b""
        if (("xml" not in ct) and (body_bytes[:1] in (b"{", b"[") or (b"<html" in body_bytes[:200].lower()))):
            self._dump_http_response(uid, r, stage="unexpected_content_type", note=f"Content-Type={ct}")
            raise RuntimeError(f"Unexpected response type for uid={uid} Content-Type={ct}")

        # Parse wrapper XML
        try:
            outer = ET.fromstring(body_bytes)
        except Exception as e:
            self._dump_http_response(uid, r, stage="outer_xml_parse_error", note=str(e))
            raise

        payload = outer.find(".//Payload")
        if payload is None or not (payload.text or "").strip():
            # sometimes API could return raw xml
            txt = r.text or ""
            if "<Dfe" in txt:
                return txt
            self._dump_http_response(uid, r, stage="no_payload", note="No <Payload> element or empty payload")
            raise RuntimeError("No Payload in DFE XML response wrapper.")

        inner = html.unescape(payload.text)
        return inner

    def _dump_http_response(self, uid: str, r: requests.Response, *, stage: str, note: str) -> None:
        """Dump raw HTTP response for offline debugging (no auth headers)."""
        try:
            dump_dir = BAD_RESPONSE_DIR or (Path.cwd() / "logs" / "bad_responses")
            dump_dir.mkdir(parents=True, exist_ok=True)
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            base = dump_dir / f"{uid}.{stage}.{ts}"

            meta_lines = [
                f"url={getattr(r, 'url', '')}",
                f"status={r.status_code}",
                f"reason={getattr(r, 'reason', '')}",
                f"note={note}",
                "headers:",
            ]
            for k, v in (r.headers or {}).items():
                meta_lines.append(f"  {k}: {v}")

            dump_text(base.with_suffix(".meta.txt"), "\n".join(meta_lines))
            # raw body
            (base.with_suffix(".body.bin")).write_bytes(r.content or b"")
            # best-effort text preview
            preview = (r.text or "")[:5000]
            dump_text(base.with_suffix(".body.txt"), preview)

            log(f"DEBUG: dumped HTTP response for uid={uid} stage={stage} -> {base}.*")
        except Exception as e:
            log(f"WARNING: failed to dump HTTP response for uid={uid}: {e}")



def extract_items(obj: Any) -> List[Dict[str, Any]]:
    """
    Attempt to normalize different possible list response shapes.
    Returns a list of dict items.
    """
    if isinstance(obj, list):
        return [x for x in obj if isinstance(x, dict)]
    if not isinstance(obj, dict):
        return []
    # common containers
    for key in ("content", "items", "data", "results", "result", "dfes", "Dfes"):
        v = obj.get(key)
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return v
    # fallback: first list-of-dicts value
    for v in obj.values():
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return v
    return []


def extract_uid_from_item(item: Dict[str, Any]) -> Optional[str]:
    candidates = [
        "Id", "ID", "Uid", "UID", "Iud", "IUD", "DfeId", "dfeId", "DocumentId", "documentId", "DocumentUid", "documentUid"
    ]
    for k in candidates:
        v = item.get(k)
        if isinstance(v, str) and UID_RE.match(v.strip()):
            return v.strip()
    # fallback: search any string field that looks like UID
    for v in item.values():
        if isinstance(v, str):
            s = v.strip()
            if UID_RE.match(s):
                return s
    return None


def extract_efatura_date_from_item(item: Dict[str, Any]) -> Optional[str]:
    """
    "Data eFatura" = authorized/registered date, best-effort.
    Returns ISO date (YYYY-MM-DD) or full datetime as string if that is what's present.
    """
    preferred_keys = [
        "AuthorizedDate", "authorizedDate", "authorized_date",
        "AuthorizedDateTime", "authorizedDateTime",
        "RegisterDate", "registerDate", "registeredDate", "registered_date",
        "CreatedAt", "createdAt", "SubmissionDate", "submissionDate",
    ]
    for k in preferred_keys:
        v = item.get(k)
        if v:
            return str(v)
    # heuristic: first key containing 'authorizeddate' or 'register'
    for k, v in item.items():
        lk = k.lower()
        if ("authorizeddate" in lk or "register" in lk) and v:
            return str(v)
    return None



DOC_PREFIX_TO_TIPO = {
    # Cabo Verde (documentos electrónicos)
    "FTE": "Fatura Eletrónica",
    "FRE": "Fatura Recibo Eletrónica",
    "TVE": "Talão de Venda Eletrónico",
    "RCE": "Recibo Eletrónico",
    "NCE": "Nota de Crédito Eletrónica",
    "NDE": "Nota de Débito Eletrónica",
    "DVE": "Nota de Devolução Eletrónica",
    "DTE": "Documento de Transporte Eletrónico",
    "NLE": "Nota de Lançamento Eletrónica",

    # Portugal / genéricos (best-effort)
    "FT": "Factura",
    "FS": "Ticket",           # Factura Simplificada
    "FR": "Factura-Recibo",
    "RC": "Recibo",
    "NC": "Nota de Crédito",
    "ND": "Nota de Débito",
    "GR": "Guia de Remessa",
    "OR": "Orçamento",
}


# DocumentTypeCode mapping (eFatura CV). The DFE root carries DocumentTypeCode and the direct child element
# indicates the document kind (Invoice, Receipt, CreditNote, etc.).
DTC_TO_META = {
    "1": {"prefix": "FTE", "label": "Fatura Eletrónica", "element": "Invoice"},
    "2": {"prefix": "FRE", "label": "Fatura Recibo Eletrónica", "element": "InvoiceReceipt"},
    "3": {"prefix": "TVE", "label": "Talão de Venda Eletrónico", "element": "SalesReceipt"},
    "4": {"prefix": "RCE", "label": "Recibo Eletrónico", "element": "Receipt"},
    "5": {"prefix": "NCE", "label": "Nota de Crédito Eletrónica", "element": "CreditNote"},
    "6": {"prefix": "NDE", "label": "Nota de Débito Eletrónica", "element": "DebitNote"},
    "7": {"prefix": "DTE", "label": "Documento de Transporte Eletrónico", "element": "Transport"},
    "8": {"prefix": "DVE", "label": "Nota de Devolução Eletrónica", "element": "ReturnNote"},
    "9": {"prefix": "NLE", "label": "Nota de Lançamento Eletrónica", "element": "RegistrationNote"},
}

DOC_ELEMENT_TO_PREFIX = {v["element"]: v["prefix"] for v in DTC_TO_META.values()}
DTC_TO_ELEMENT = {k: v["element"] for k, v in DTC_TO_META.items()}


def infer_tipo_documento(document_number: str, doc_kind: str = "") -> str:
    """Infer 'Tipo de Documento' using DocumentNumber prefix and/or doc_kind (best-effort)."""
    num = (document_number or "").strip()
    if num:
        m = re.match(r"^\s*([A-Za-z]{1,4})\b", num)
        if m:
            pref = m.group(1).upper()
            if pref in DOC_PREFIX_TO_TIPO:
                return DOC_PREFIX_TO_TIPO[pref]
    kind = (doc_kind or "").strip().lower()
    if kind == "invoice":
        return "Factura"
    if kind == "receipt":
        return "Recibo"
    if doc_kind:
        return doc_kind
    return ""

def safe_save_workbook(wb: Workbook, path: Path) -> None:
    """Write to temp then replace for improved crash safety."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)



def delete_uid_rows(ws: Worksheet, uid: str) -> int:
    """Delete all Excel rows (excluding header) for a given UID. Returns number of deleted rows."""
    uid = (uid or "").strip()
    if not uid:
        return 0
    deleted = 0
    # delete bottom-up to avoid index shift issues
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == uid:
            ws.delete_rows(r, 1)
            deleted += 1
    return deleted


def _read_header(ws: Worksheet) -> List[str]:
    vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    out: List[str] = []
    for v in vals:
        if v is None:
            out.append("")
        else:
            out.append(str(v).strip())
    # trim trailing empty
    while out and out[-1] == "":
        out.pop()
    return out



# =============================================================================
# Excel I/O and UID re-write mechanics
# =============================================================================

def ensure_workbook(path: Path) -> Tuple[Workbook, Worksheet, Dict[str, List[int]]]:
    """
    Returns wb, ws, uid_row_map (UID -> list of row indices with that UID)

    If an existing workbook matches the previous schema (COLUMNS_V1), it will be
    migrated in-place to the current schema (COLUMNS) by inserting the
    "Tipo de Documento" column between "Data Documento" and "Numero Documento".
    """
    uid_map: Dict[str, List[int]] = {}

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active

        header = _read_header(ws)
        if header == COLUMNS:
            pass
        elif header == COLUMNS_V1:
            log(f"Excel header v1 detected in {path}; migrating to v2 (adding 'Tipo de Documento').")
            insert_at = header.index("Numero Documento") + 1  # 1-based, insert BEFORE Numero Documento
            ws.insert_cols(insert_at)
            ws.cell(row=1, column=insert_at, value="Tipo de Documento")

            # enforce full header (guards against accidental shifts)
            for c, name in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=c, value=name)

            # best-effort backfill
            col_docnum = COLUMNS.index("Numero Documento") + 1
            col_tipo = COLUMNS.index("Tipo de Documento") + 1
            for r in range(2, ws.max_row + 1):
                cur = ws.cell(row=r, column=col_tipo).value
                if cur is None or str(cur).strip() == "":
                    docnum = ws.cell(row=r, column=col_docnum).value
                    tipo = infer_tipo_documento(str(docnum) if docnum is not None else "")
                    if tipo:
                        ws.cell(row=r, column=col_tipo, value=tipo)

            safe_save_workbook(wb, path)
            log("Excel migration saved.")
        else:
            raise RuntimeError(
                f"Excel header mismatch in {path}. Expected columns: {COLUMNS} (or legacy {COLUMNS_V1}), got: {header}"
            )

        for r in range(2, ws.max_row + 1):
            uid = ws.cell(row=r, column=1).value
            if isinstance(uid, str) and uid.strip():
                uid_map.setdefault(uid.strip(), []).append(r)
        return wb, ws, uid_map

    # create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "supplier_invoices"
    for c, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=c, value=name)
    ws.freeze_panes = "A2"
    return wb, ws, uid_map



def parse_supplier_address(party_el: Optional[ET.Element]) -> str:
    """Best-effort address extraction (namespace-agnostic).

    Real-world eFatura CV payloads are not always consistent about namespaces, so we avoid d:Address queries.
    """
    if party_el is None:
        return ""

    addr = _find_first_by_localnames(party_el, ["Address"])
    if addr is None:
        return ""

    def _child_text(el: ET.Element, local: str) -> str:
        for ch in list(el):
            if _localname(ch.tag) == local:
                return get_text(ch)
        return ""

    parts: List[str] = []
    for tag in ("Street", "BuildingFloor", "AddressDetail", "City", "PostalCode"):
        t = _child_text(addr, tag)
        if t:
            parts.append(t)
    return ", ".join(parts)


def _pick_document_node(dfe_root: ET.Element) -> ET.Element:
    """Pick the *primary* fiscal document node from a DFE root.

    IMPORTANT: Many DFEs contain <FiscalDocument> references *inside* the document body.
    A naive deep-search will often pick a reference node first, leading to 'no lines' false positives.
    The schema defines the document kind as a direct child of <Dfe>.
    """
    dtc = (dfe_root.attrib.get("DocumentTypeCode") or "").strip()
    expected = DTC_TO_ELEMENT.get(dtc)

    # First pass: direct child that matches DocumentTypeCode expectation
    for ch in list(dfe_root):
        ln = _localname(ch.tag)
        if ln in ("Signature", "SignedInfo", "KeyInfo"):
            continue
        if expected and ln == expected:
            return ch

    # Second pass: any direct child that matches known document elements
    for ch in list(dfe_root):
        ln = _localname(ch.tag)
        if ln in DOC_ELEMENT_TO_PREFIX:
            return ch

    # Last resort: fallback to deep search
    found = _find_first_by_localnames(dfe_root, list(DOC_ELEMENT_TO_PREFIX.keys()))
    return found if found is not None else dfe_root




def parse_invoice_lines(dfe_root: ET.Element) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Parse a DFE XML into (meta, lines), tolerating multiple document kinds.

    Key points (from observed eFatura CV payloads):
    - The DFE root has attribute DocumentTypeCode and the *primary* document is a direct child element.
    - Many documents include <FiscalDocument> references which must not be mistaken for the primary doc node.
    - Receipts (RCE / DocumentTypeCode=4) may contain no lines and only reference a FiscalDocument UID.
    """
    doc_meta: Dict[str, Any] = {
        "supplier_name": "",
        "supplier_taxid": "",
        "supplier_address": "",
        "issue_date": "",
        "document_number": "",
        "doc_kind": "",  # should be one of FTE/FRE/TVE/RCE/NCE/NDE/DVE/DTE/NLE when possible
        "doc_kind_label": "",
        "document_type_code": (dfe_root.attrib.get("DocumentTypeCode") or "").strip(),
        "ref_uid": "",
        "ref_uids": [],
        "doc_node": "",
    }

    dtc = doc_meta["document_type_code"]
    if dtc in DTC_TO_META:
        doc_meta["doc_kind"] = DTC_TO_META[dtc]["prefix"]
        doc_meta["doc_kind_label"] = DTC_TO_META[dtc]["label"]

    # identify primary document node (direct child)
    doc_node = _pick_document_node(dfe_root)
    doc_node_name = _localname(doc_node.tag)
    doc_meta["doc_node"] = doc_node_name

    # If we still don't have a kind, infer from element name
    if not doc_meta["doc_kind"]:
        doc_meta["doc_kind"] = DOC_ELEMENT_TO_PREFIX.get(doc_node_name, doc_node_name or "Unknown")
        doc_meta["doc_kind_label"] = DOC_PREFIX_TO_TIPO.get(doc_meta["doc_kind"], "")

    # emitter/supplier info (try under doc_node; fallback anywhere)
    emitter = None
    for cand in ("EmitterParty", "SellerParty", "SupplierParty", "AccountingSupplierParty"):
        emitter = _find_first_by_localnames(doc_node, [cand])
        if emitter is not None:
            break
    if emitter is None:
        emitter = _find_first_by_localnames(dfe_root, ["EmitterParty"])

    if emitter is not None:
        doc_meta["supplier_name"] = _coalesce(
            _text_anywhere(emitter, ["Name", "PartyName"]),
            _text_anywhere(doc_node, ["EmitterName", "SupplierName"]),
        )
        doc_meta["supplier_taxid"] = _coalesce(
            _text_anywhere(emitter, ["TaxId", "TaxID", "CompanyID", "VatID"]),
            _text_anywhere(doc_node, ["TaxId", "TaxID"]),
        )
        doc_meta["supplier_address"] = parse_supplier_address(emitter)

    # issue date and document number
    doc_meta["issue_date"] = _coalesce(
        _text_anywhere(doc_node, ["IssueDate", "IssueDateTime", "AuthorizedDateTime"]),
        _text_anywhere(dfe_root, ["IssueDate", "IssueDateTime", "AuthorizedDateTime"]),
    )

    # some documents split Serie + DocumentNumber
    serie = _text_anywhere(doc_node, ["Serie"])
    docnum = _text_anywhere(doc_node, ["DocumentNumber"])
    if serie and docnum:
        doc_meta["document_number"] = f"{serie}/{docnum}"
    else:
        doc_meta["document_number"] = _coalesce(
            docnum,
            serie,
            _text_anywhere(doc_node, ["Number", "DocumentId", "DocumentID", "ID"]),
        )

    # references (for receipts and related docs)
    refs = _find_reference_uids(doc_node)
    if refs:
        doc_meta["ref_uids"] = refs
        doc_meta["ref_uid"] = refs[0]

    # lines: some payloads contain multiple <Lines> nodes (e.g., references/embedded structures).
    # We scan all candidate <Lines> elements and take the first that yields actual <...Line> items.
    lines: List[Dict[str, Any]] = []

    def _scan_lines(root: ET.Element) -> List[Dict[str, Any]]:
        for el in root.iter():
            if _localname(el.tag) != "Lines":
                continue
            parsed = parse_lines(el)
            if parsed:
                return parsed
        return []

    lines = _scan_lines(doc_node)
    if not lines and doc_node is not dfe_root:
        lines = _scan_lines(dfe_root)

    return doc_meta, lines

def parse_lines(lines_el: Optional[ET.Element]) -> List[Dict[str, Any]]:
    """Parse line items from a <Lines> element.

    eFatura CV typically uses <Lines><Line>..., but in practice we may see:
    - different namespaces
    - alternative node names (e.g. InvoiceLine/CreditNoteLine)
    This function is tolerant and uses localname-based fallbacks.
    """
    if lines_el is None:
        return []
    out: List[Dict[str, Any]] = []

    # Preferred: DFE namespace Line
    candidates = list(lines_el.findall("d:Line", NS_DFE))

    # Fallback: any direct children whose localname endswith 'Line'
    if not candidates:
        for ch in list(lines_el):
            if _localname(ch.tag).lower().endswith("line"):
                candidates.append(ch)

    # Last resort: any descendant node whose localname endswith 'Line' (can overmatch but better than empty)
    if not candidates:
        for ch in lines_el.iter():
            if ch is lines_el:
                continue
            if _localname(ch.tag).lower().endswith("line"):
                candidates.append(ch)

    for line in candidates:
        # qty/unit
        qty_el = _find_first_by_localnames(line, ["Quantity", "InvoicedQuantity", "CreditedQuantity", "DebitedQuantity"])
        qty = safe_float(get_text(qty_el))
        unit = ""
        if qty_el is not None:
            unit = (qty_el.attrib.get("UnitCode") or qty_el.attrib.get("unitCode") or "").strip()

        # prices/totals
        unit_price = safe_float(_text_anywhere(line, ["Price", "UnitPrice", "PriceAmount"]))
        ext = safe_float(_text_anywhere(line, ["PriceExtension", "LineExtensionAmount"]))
        net = safe_float(_text_anywhere(line, ["NetTotal", "LineTotal"]))
        total = safe_float(_text_anywhere(line, ["Total", "Amount"]))

        # item details
        item = _find_first_by_localnames(line, ["Item", "Product", "GoodsItem"])
        item_name = ""
        item_code = ""
        if item is not None:
            item_name = _coalesce(
                _text_anywhere(item, ["Description", "Name", "ItemName"]),
                _text_anywhere(line, ["Description", "Name"]),
            )
            item_code = _coalesce(
                _text_anywhere(item, ["EmitterIdentification", "SellerItemIdentification", "ID", "Code"]),
                _text_anywhere(line, ["EmitterIdentification", "SellerItemIdentification", "ID", "Code"]),
            )
        else:
            item_name = _text_anywhere(line, ["Description", "Name"])
            item_code = _text_anywhere(line, ["EmitterIdentification", "SellerItemIdentification", "ID", "Code"])

        # discount: explicit fields if present; else compute diff where possible
        discount = safe_float(_text_anywhere(line, ["Discount", "DiscountAmount"]))
        if discount is None:
            if ext is not None and qty is not None and unit_price is not None:
                discount = round(max(unit_price * qty - ext, 0.0), 2)
            elif net is not None and qty is not None and unit_price is not None:
                discount = round(max(unit_price * qty - net, 0.0), 2)

        # choose line_total
        line_total = None
        for v in (net, ext, total):
            if v is not None:
                line_total = v
                break
        if line_total is None and qty is not None and unit_price is not None:
            line_total = round(qty * unit_price, 2)

        out.append(
            {
                "item_code": item_code,
                "item_name": item_name,
                "qty": qty,
                "unit": unit,
                "unit_price": unit_price,
                "discount": discount,
                "line_total": line_total,
            }
        )
    return out

def append_error_row(ws: Worksheet, uid: str, reason: str) -> None:
    col = {name: (COLUMNS.index(name) + 1) for name in COLUMNS}
    r = ws.max_row + 1
    ws.cell(row=r, column=col["UID"], value=uid)
    ws.cell(row=r, column=col["Erro"], value=reason)
    ws.cell(row=r, column=col["last_updated"], value=now_local_iso())



def append_line_rows(ws: Worksheet, uid: str, efatura_date: str, meta: Dict[str, Any], lines: List[Dict[str, Any]]) -> int:
    """
    Append one row per line item. Returns number of rows added.
    """
    col = {name: (COLUMNS.index(name) + 1) for name in COLUMNS}
    tipo_doc = infer_tipo_documento(meta.get("document_number", ""), meta.get("doc_kind", ""))

    before = ws.max_row
    for ln in lines:
        r = ws.max_row + 1
        ws.cell(row=r, column=col["UID"], value=uid)
        ws.cell(row=r, column=col["Erro"], value="")
        ws.cell(row=r, column=col["Nome Fornecedor"], value=meta.get("supplier_name", ""))
        ws.cell(row=r, column=col["NIF Fornecedor"], value=meta.get("supplier_taxid", ""))
        ws.cell(row=r, column=col["Morada Fornecedor"], value=meta.get("supplier_address", ""))
        ws.cell(row=r, column=col["Data eFatura"], value=efatura_date or "")
        ws.cell(row=r, column=col["Data Documento"], value=meta.get("issue_date", ""))
        ws.cell(row=r, column=col["Tipo de Documento"], value=tipo_doc)
        ws.cell(row=r, column=col["Numero Documento"], value=meta.get("document_number", ""))
        ws.cell(row=r, column=col["Código Artigo"], value=ln.get("item_code", ""))
        ws.cell(row=r, column=col["Nome Artigo"], value=ln.get("item_name", ""))
        ws.cell(row=r, column=col["Quantidade"], value=ln.get("qty"))
        ws.cell(row=r, column=col["Unidade Medida"], value=ln.get("unit", ""))
        ws.cell(row=r, column=col["Preço Unitário"], value=ln.get("unit_price"))
        ws.cell(row=r, column=col["Desconto"], value=ln.get("discount"))
        ws.cell(row=r, column=col["Preço Total (linha)"], value=ln.get("line_total"))
        ws.cell(row=r, column=col["last_updated"], value=now_local_iso())
    return ws.max_row - before



def backfill_efatura_dates(ws: Worksheet, uid_row_map: Dict[str, List[int]], uid_to_efdate: Dict[str, str]) -> int:
    """
    Fill 'Data eFatura' for existing rows when empty.
    Returns number of rows updated.
    """
    col_ef = COLUMNS.index("Data eFatura") + 1
    col_last = COLUMNS.index("last_updated") + 1
    updated = 0
    for uid, rows in uid_row_map.items():
        ef = uid_to_efdate.get(uid)
        if not ef:
            continue
        for r in rows:
            cur = ws.cell(row=r, column=col_ef).value
            if cur is None or str(cur).strip() == "":
                ws.cell(row=r, column=col_ef, value=ef)
                ws.cell(row=r, column=col_last, value=now_local_iso())
                updated += 1
    return updated


def load_token_json(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"token.json not found: {path}")
    obj = json.loads(path.read_text(encoding="utf-8"))
    tok = obj.get("access_token")
    if not tok:
        raise ValueError("token.json missing access_token")
    return str(tok)



# =============================================================================
# Main entrypoint
# =============================================================================

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="Path to INI file")
    ap.add_argument("--show-fields", action="store_true", help="Show listing response fields and exit")
    ap.add_argument("--verbose", action="store_true", help="Verbose HTTP logging")
    ap.add_argument("--max-docs", type=int, default=0, help="Process at most N documents (0 = no limit)")
    ap.add_argument("--save-every-docs", type=int, default=-1, help="Save Excel every N processed docs (0 disables; -1 uses INI)")
    ap.add_argument("--save-every-seconds", type=int, default=-1, help="Save Excel every N seconds (0 disables; -1 uses INI)")
    ap.add_argument("--log-file", default="", help="Path to log file (default: base_dir/logs/update_supplier_invoices_YYYYmmdd_HHMMSS.log)")
    ap.add_argument("--rewrite-existing", action="store_true", help="If UID already exists in Excel, delete its rows and rewrite it (WARNING: can re-download many docs if used broadly).")
    args = ap.parse_args()

    cfg = load_config(Path(args.config).expanduser(), verbose=args.verbose)

    if args.log_file:
        # If relative, resolve relative to base_dir
        lf = args.log_file.strip()
        cfg.log_file = (cfg.base_dir / lf).expanduser().resolve() if not os.path.isabs(lf) else Path(lf).expanduser().resolve()

    setup_logging(cfg.log_file)
    log(f"Log file: {cfg.log_file}")

    global BAD_RESPONSE_DIR
    BAD_RESPONSE_DIR = (cfg.log_file.parent / "bad_responses").resolve()
    BAD_RESPONSE_DIR.mkdir(parents=True, exist_ok=True)

    if args.save_every_docs != -1:
        cfg.save_every_docs = max(0, int(args.save_every_docs))
    if args.save_every_seconds != -1:
        cfg.save_every_seconds = max(0, int(args.save_every_seconds))

    # Excel path and base_dir
    log(f"Base dir: {cfg.base_dir}")
    log(f"Excel: {cfg.excel_path}")

    # DNS preflight
    try:
        ips_services = resolve_or_fail("services.efatura.cv")
        log(f"DNS OK: services.efatura.cv -> {', '.join(ips_services)}")
    except Exception as e:
        log(f"ERROR: {e}")
        return 2
    try:
        ips_iam = resolve_or_fail("iam.efatura.cv")
        log(f"DNS OK: iam.efatura.cv -> {', '.join(ips_iam)}")
    except Exception as e:
        log(f"WARNING: {e} (userinfo/refresh may fail)")
        # do not hard-fail; XML fetch/list only needs services.efatura.cv

    auth = EfaturaAuthManager(
        issuer_url=cfg.auth_issuer_url,
        client_id=cfg.auth_client_id,
        redirect_uri=cfg.auth_redirect_uri,
        scopes=cfg.auth_scopes,
        token_store_path=cfg.auth_token_store,
        timeout=cfg.timeout_sec,
        retries=cfg.retries,
        client_secret=cfg.auth_client_secret,
    )
    if auth.migrate_legacy_tokens(cfg.token_json):
        log(f"Migrated legacy token.json to {cfg.auth_token_store}")

    try:
        access_token_provider = lambda: auth.get_valid_access_token()
        _ = access_token_provider()
    except EfaturaAuthNeedsReauth:
        log("ERROR: Autenticação eFatura necessária. Abra Configurações > eFatura > Ligar Conta para concluir o login.")
        return 3

    client = EfaturaClient(
        access_token_provider=access_token_provider,
        repo_code=cfg.repo_code,
        timeout_sec=cfg.timeout_sec,
        retries=cfg.retries,
        backoff_sec=cfg.backoff_sec,
        verbose=args.verbose,
    )

    # userinfo (best effort)
    try:
        taxid = client.userinfo_taxid()
        log(f"eFatura userinfo OK: {taxid}")
    except EfaturaAuthNeedsReauth:
        log("ERROR: Autenticação eFatura necessária. Abra Configurações > eFatura > Ligar Conta para concluir o login.")
        return 3
    except PermissionError:
        log("ERROR: TOKEN_EXPIRED_OR_INVALID (userinfo).")
        return 3
    except Exception as e:
        log(f"WARNING: userinfo failed: {e} (continuing)")

    wb, ws, uid_row_map = ensure_workbook(cfg.excel_path)
    existing_uids = set(uid_row_map.keys())
    log(f"Existing UIDs in Excel: {len(existing_uids)}")

    # Resume safety: if we crashed mid-document, rewrite that UID (delete existing rows + re-fetch).
    resume_state_path = cfg.excel_path.with_suffix(cfg.excel_path.suffix + ".resume.json")
    resume_state = load_resume_state(resume_state_path)
    resume_uid = compute_resume_uid(resume_state)
    if resume_uid:
        log(f"Resume detected: last started UID={resume_uid} was not marked completed. Will rewrite it.")
    else:
        log("Resume state: clean (no in-progress UID).")

    log(f"Listing DFEs from {cfg.date_start} to {cfg.date_end} (repo={cfg.repo_code}, page_size={cfg.page_size})")
    try:
        records, discovered_date_keys = client.list_dfes(cfg.date_start, cfg.date_end, cfg.page_size, show_fields=args.show_fields)
    except EfaturaAuthNeedsReauth:
        log("ERROR: Autenticação eFatura necessária. Abra Configurações > eFatura > Ligar Conta para concluir o login.")
        return 3
    except PermissionError:
        log("ERROR: TOKEN_EXPIRED_OR_INVALID (listing).")
        return 3

    if args.show_fields:
        return 0

    if discovered_date_keys:
        log(f"Discovered possible 'Data eFatura' fields in listing: {', '.join(discovered_date_keys)}")

    uid_to_item: Dict[str, Dict[str, Any]] = {}
    uid_to_efdate: Dict[str, str] = {}
    for it in records:
        uid = extract_uid_from_item(it)
        if not uid:
            continue
        uid_to_item[uid] = it
        efdate = extract_efatura_date_from_item(it)
        if efdate:
            uid_to_efdate[uid] = efdate

    # backfill Data eFatura for existing docs
    filled = backfill_efatura_dates(ws, uid_row_map, uid_to_efdate)
    if filled:
        log(f"Backfilled 'Data eFatura' for {filled} existing rows.")

    uids = sorted(uid_to_item.keys())
    log(f"Total UIDs discovered in date range: {len(uids)}")

    added_docs = 0
    added_rows = 0
    errors = 0

    last_save_ts = time.time()
    docs_since_save = 0

    def checkpoint_save(force: bool = False) -> None:
        nonlocal last_save_ts, docs_since_save
        if force:
            safe_save_workbook(wb, cfg.excel_path)
            last_save_ts = time.time()
            docs_since_save = 0
            log(f"Excel checkpoint saved: {cfg.excel_path}")
            return

        if cfg.save_every_docs == 0 and cfg.save_every_seconds == 0:
            return

        due_by_docs = cfg.save_every_docs > 0 and docs_since_save >= cfg.save_every_docs
        due_by_time = cfg.save_every_seconds > 0 and (time.time() - last_save_ts) >= cfg.save_every_seconds
        if due_by_docs or due_by_time:
            safe_save_workbook(wb, cfg.excel_path)
            last_save_ts = time.time()
            docs_since_save = 0
            log(f"Excel checkpoint saved: {cfg.excel_path}")


    for idx, uid in enumerate(uids, start=1):
        if args.max_docs and added_docs >= args.max_docs:
            log(f"Reached --max-docs={args.max_docs}, stopping.")
            break

        uid_exists = uid in existing_uids
        should_rewrite = args.rewrite_existing or (resume_uid == uid)

        if uid_exists and not should_rewrite:
            if idx % max(1, cfg.progress_every) == 0:
                log(f"[{idx}/{len(uids)}] UID={uid} already in Excel -> skip")
            continue

        if uid_exists and should_rewrite:
            deleted = delete_uid_rows(ws, uid)
            if deleted:
                log(f"UID={uid} already existed in Excel: deleted {deleted} row(s) to rewrite document.")
            # Keep existing_uids set; UID still considered known.

        # Mark started (resume checkpoint)
        resume_state["started_uid"] = uid
        save_resume_state(resume_state_path, resume_state)

        log(f"[{idx}/{len(uids)}] UID={uid} fetching document XML...")
        efdate = uid_to_efdate.get(uid, "")

        try:
            inner_xml = client.fetch_dfe_inner_xml(uid)
            dfe_root = safe_parse_xml(inner_xml, uid=uid, stage="inner", dump_dir=(BAD_RESPONSE_DIR or (cfg.log_file.parent / "bad_responses")))
            meta, lines = parse_invoice_lines(dfe_root)

            # If we have no lines, try to follow referenced FiscalDocument UIDs (receipts, notes, etc.)
            if not lines:
                ref_uids = meta.get("ref_uids") or ([meta["ref_uid"]] if meta.get("ref_uid") else [])
                ref_uids = [r for r in ref_uids if r and r != uid]

                for ref_uid in ref_uids:
                    log(f"UID={uid} has no lines; trying referenced FiscalDocument {ref_uid}...")
                    try:
                        inner2 = client.fetch_dfe_inner_xml(ref_uid)
                        root2 = safe_parse_xml(inner2, uid=ref_uid, stage="inner_ref", dump_dir=(BAD_RESPONSE_DIR or (cfg.log_file.parent / "bad_responses")))
                        meta2, lines2 = parse_invoice_lines(root2)
                    except Exception as e:
                        log(f"WARNING: referenced fetch/parse failed ref_uid={ref_uid}: {e}")
                        continue

                    if lines2:
                        # Keep original document identity (number/type/date) but backfill supplier fields if missing.
                        meta = dict(meta)  # copy
                        for k in ("supplier_name", "supplier_taxid", "supplier_address"):
                            if not meta.get(k):
                                meta[k] = meta2.get(k, "")
                        lines = lines2
                        break
            if not lines:
                # Dump the raw inner xml for forensic analysis (helps identify unexpected schemas)
                try:
                    nl_dir = cfg.log_file.parent / "no_lines"
                    nl_dir.mkdir(parents=True, exist_ok=True)
                    dump_text(nl_dir / f"{uid}.inner.xml", inner_xml)
                except Exception as _e:
                    log(f"WARNING: failed to dump no-lines XML for uid={uid}: {_e}")

                doc_kind = meta.get("doc_kind") or ""
                doc_num = meta.get("document_number") or ""
                refs = meta.get("ref_uids") or ([meta["ref_uid"]] if meta.get("ref_uid") else [])
                refs_s = ",".join(refs) if refs else ""

                append_error_row(
                    ws,
                    uid,
                    f"No Lines found (doc_kind={doc_kind}, doc_number={doc_num}, refs={refs_s})",
                )
                errors += 1
                existing_uids.add(uid)
                uid_row_map.setdefault(uid, []).append(ws.max_row)
                log(f"WARNING: UID={uid} has no lines; recorded as error. kind={doc_kind} num={doc_num} refs={refs_s}")

                # Mark completed (resume checkpoint)
                resume_state["completed_uid"] = uid
                save_resume_state(resume_state_path, resume_state)
                docs_since_save += 1
                checkpoint_save()
                continue

            before = ws.max_row
            rows_added = append_line_rows(ws, uid, efdate, meta, lines)
            after = ws.max_row
            added_rows += rows_added
            added_docs += 1
            existing_uids.add(uid)
            # record new uid rows for future backfill
            uid_row_map[uid] = list(range(before + 1, after + 1))

            # Mark completed (resume checkpoint)
            resume_state["completed_uid"] = uid
            save_resume_state(resume_state_path, resume_state)
            if resume_uid == uid:
                resume_uid = None

            log(f"OK UID={uid} lines={len(lines)} supplier={meta.get('supplier_name','')}")
            docs_since_save += 1
            checkpoint_save()
        except EfaturaAuthNeedsReauth:
            log("ERROR: Autenticação eFatura necessária. Abra Configurações > eFatura > Ligar Conta para concluir o login.")
            break
        except PermissionError:
            log("ERROR: TOKEN_EXPIRED_OR_INVALID while fetching documents. Stop now.")
            break
        except Exception as e:
            append_error_row(ws, uid, str(e)[:500])
            errors += 1
            existing_uids.add(uid)
            uid_row_map.setdefault(uid, []).append(ws.max_row)
            log_exception(f"WARNING: Failed UID={uid}: {e}")

            # Mark completed (resume checkpoint)
            resume_state["completed_uid"] = uid
            save_resume_state(resume_state_path, resume_state)
            if resume_uid == uid:
                resume_uid = None
            docs_since_save += 1
            checkpoint_save()

        if (idx % max(1, cfg.progress_every)) == 0:
            log(f"Progress: processed={idx}/{len(uids)} added_docs={added_docs} added_rows={added_rows} errors={errors}")

    checkpoint_save(force=True)
    log(f"DONE. Added docs={added_docs}, rows={added_rows}, errors={errors}. Excel saved: {cfg.excel_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        log("Interrupted by user.")
        raise
